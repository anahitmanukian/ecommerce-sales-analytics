import os
import random
from datetime import datetime, timedelta
import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

# Set seeds for reproducibility
random.seed(42)
np.random.seed(42)

# -----------------------------
# 1. Config & Business Setup
# -----------------------------
n_customers = 3500
n_products = 180
n_orders = 22000
n_order_items = 30000
n_returns = 2600

first_names = ["Ani", "Mariam", "Anna", "Arman", "Narek", "Hayk", "Lilit", "Sona", "David", "Gor", "Mane", "Suren", "Tatev", "Arsen", "Eva", "Mikael", "Lusine", "Karen", "Saro", "Nare"]
last_names = ["Petrosyan", "Hakobyan", "Grigoryan", "Sargsyan", "Manukyan", "Harutyunyan", "Martirosyan", "Karapetyan", "Mkrtchyan", "Avetisyan", "Hovhannisyan", "Ghazaryan", "Stepanyan", "Vardanyan"]
cities = ["Yerevan", "Gyumri", "Vanadzor", "Abovyan", "Hrazdan", "Etchmiadzin", "Kapan", "Armavir", "Artashat", "Dilijan"]
categories = ["Electronics", "Home", "Beauty", "Sports", "Clothing", "Books", "Accessories"]
brands = ["Ararat Tech", "Nova", "UrbanLife", "Peak", "Mira", "Vertex", "Astra", "HomePro", "Lumen", "Nexo"]
payment_methods = ["Card", "Cash", "Bank Transfer", "PayPal", "Apple Pay"]
channels = ["Website", "Mobile App", "Marketplace", "Social Media"]
statuses = ["Completed", "Completed", "Completed", "Completed", "Cancelled", "Pending", "Shipped"]

city_abbrevs = {
    "Yerevan": ["Yer", "Yerev.", "YVN", "YEREVAN", "yerevan city", "Yev."],
    "Gyumri": ["Gym", "Gyumri City", "GMR", "gyumri"],
    "Vanadzor": ["Vandz.", "VNZ", "vanadzor"],
    "Etchmiadzin": ["Vagharshapat", "Etch.", "Etchmiadzin"],
}

# -----------------------------
# 2. Noise Generation Helper
# -----------------------------
def corrupt_value(val, abbrev_map=None, p_punct=0.30, p_case=0.35, p_abbrev=0.40):
    if pd.isna(val):
        return val
    s = str(val)
    
    # Regional Abbreviations / Typos
    if abbrev_map and s in abbrev_map and random.random() < p_abbrev:
        s = random.choice(abbrev_map[s])
        
    # Mixed Casing
    if random.random() < p_case:
        choice = random.choice(["lower", "upper", "mixed"])
        if choice == "lower":
            s = s.lower()
        elif choice == "upper":
            s = s.upper()
        else:
            s = "".join(c.upper() if random.random() > 0.5 else c.lower() for c in s)
            
    # Quotes, Commas, Leading/Trailing Whitespace
    if random.random() < p_punct:
        wrappers = [f'"{s}"', f"'{s}'", f"  {s}  ", f"{s},", f"#{s}", f"{s}.", f'""{s}""']
        s = random.choice(wrappers)
        
    return s

# -----------------------------
# 3. Data Generation
# -----------------------------
# Customers
customer_ids = [f"C{10000+i}" for i in range(n_customers)]
customers = pd.DataFrame({
    "customer_id": customer_ids,
    "first_name": np.random.choice(first_names, n_customers),
    "last_name": np.random.choice(last_names, n_customers),
    "email": [f"customer{i}@example.com" for i in range(n_customers)],
    "city": np.random.choice(cities, n_customers, p=[.45, .10, .09, .07, .06, .06, .04, .05, .04, .04]),
    "signup_date": [datetime(2023, 1, 1) + timedelta(days=random.randint(0, 950)) for _ in range(n_customers)],
    "age": np.random.randint(18, 66, n_customers).astype(object),
})

# Products
product_ids = [f"P{1000+i}" for i in range(n_products)]
products = pd.DataFrame({
    "product_id": product_ids,
    "product_name": [f"{random.choice(brands)} {random.choice(['Pro','Plus','Lite','Max','Classic','X','Go'])} {i+1}" for i in range(n_products)],
    "category": np.random.choice(categories, n_products),
    "brand": np.random.choice(brands, n_products),
    "unit_cost": np.round(np.random.uniform(5, 180, n_products), 2),
})
products["list_price"] = np.round(products["unit_cost"] * np.random.uniform(1.25, 2.7, n_products), 2)

# Orders
order_ids = [f"O{100000+i}" for i in range(n_orders)]
orders = pd.DataFrame({
    "order_id": order_ids,
    "customer_id": np.random.choice(customer_ids, n_orders),
    "order_date": [datetime(2025, 1, 1) + timedelta(days=random.randint(0, 364)) for _ in range(n_orders)],
    "payment_method": np.random.choice(payment_methods, n_orders, p=[.48, .12, .16, .16, .08]),
    "sales_channel": np.random.choice(channels, n_orders, p=[.50, .27, .15, .08]),
    "order_status": np.random.choice(statuses, n_orders),
    "shipping_city": np.random.choice(cities, n_orders),
    "shipping_fee": np.round(np.random.choice([0, 0, 0, 2.5, 4.5, 6.0, 8.5], n_orders), 2)
})

# Order Items
item_ids = [f"OI{200000+i}" for i in range(n_order_items)]
order_items = pd.DataFrame({
    "order_item_id": item_ids,
    "order_id": np.random.choice(order_ids, n_order_items),
    "product_id": np.random.choice(product_ids, n_order_items),
    "quantity": np.random.choice([1, 1, 1, 2, 2, 3, 4, 5], n_order_items).astype(object),
    "unit_price": 0.0,
    "discount_pct": np.random.choice([0, .05, .10, .15, .20, .25], n_order_items, p=[.40, .15, .20, .12, .08, .05])
})
price_lookup = products.set_index("product_id")["list_price"].to_dict()
order_items["unit_price"] = order_items["product_id"].map(price_lookup).fillna(0)
order_items["unit_price"] = np.round(order_items["unit_price"] * np.random.uniform(.92, 1.05, n_order_items), 2)

# Returns
return_ids = [f"R{30000+i}" for i in range(n_returns)]
returns = pd.DataFrame({
    "return_id": return_ids,
    "order_id": np.random.choice(order_ids, n_returns),
    "return_date": [datetime(2025, 1, 5) + timedelta(days=random.randint(3, 180)) for _ in range(n_returns)],
    "return_reason": np.random.choice(["Damaged", "Wrong Item", "Changed Mind", "Late Delivery", "Not as Described", "Size Issue"], n_returns),
    "refund_amount": np.round(np.random.uniform(5, 350, n_returns), 2)
})

# -----------------------------
# 4. Apply In-Place Messiness
# -----------------------------
# Apply String Punctuation/Abbreviations/Casing
customers["city"] = customers["city"].apply(lambda x: corrupt_value(x, abbrev_map=city_abbrevs))
customers["first_name"] = customers["first_name"].apply(lambda x: corrupt_value(x))
customers["last_name"] = customers["last_name"].apply(lambda x: corrupt_value(x))
products["category"] = products["category"].apply(lambda x: corrupt_value(x))
orders["shipping_city"] = orders["shipping_city"].apply(lambda x: corrupt_value(x, abbrev_map=city_abbrevs))
orders["payment_method"] = orders["payment_method"].apply(lambda x: corrupt_value(x))
returns["return_reason"] = returns["return_reason"].apply(lambda x: corrupt_value(x))

# Mixed Data Types & Numeric Anomaly Injection
customers.loc[np.random.choice(customers.index, 120, replace=False), "email"] = np.nan
customers.loc[np.random.choice(customers.index, 80, replace=False), "age"] = "N/A"
customers.loc[np.random.choice(customers.index, 40, replace=False), "age"] = "-5"

products.loc[np.random.choice(products.index, 15, replace=False), "list_price"] *= -1
products.loc[np.random.choice(products.index, 10, replace=False), "unit_cost"] = np.nan

orders.loc[np.random.choice(orders.index, 80, replace=False), "customer_id"] = "UNKNOWN"
orders.loc[np.random.choice(orders.index, 120, replace=False), "customer_id"] = np.nan

bad_q = np.random.choice(order_items.index, 200, replace=False)
order_items.loc[bad_q[:60], "quantity"] = 0
order_items.loc[bad_q[60:100], "quantity"] = "-1"
order_items.loc[bad_q[100:150], "quantity"] = "MISSING"


# Date Formatting Corruption
date_idx = np.random.choice(orders.index, 400, replace=False)

# Convert specific subsets directly to formatted strings without casting the entire column to object beforehand
orders.loc[date_idx[:100], "order_date"] = orders.loc[date_idx[:100], "order_date"].dt.strftime("%d/%m/%Y")
orders.loc[date_idx[100:200], "order_date"] = orders.loc[date_idx[100:200], "order_date"].dt.strftime("%m-%d-%Y")
orders.loc[date_idx[200:300], "order_date"] = orders.loc[date_idx[200:300], "order_date"].dt.strftime("%b %d, %Y")

# Cast to object or string to allow inserting invalid date strings
orders["order_date"] = orders["order_date"].astype(object)
orders.loc[date_idx[300:], "order_date"] = "2025/13/45"  # Unparseable

# Duplicate Injection
customers = pd.concat([customers, customers.sample(55, random_state=7)], ignore_index=True)
orders = pd.concat([orders, orders.sample(120, random_state=12)], ignore_index=True)
order_items = pd.concat([order_items, order_items.sample(140, random_state=22)], ignore_index=True)
returns = pd.concat([returns, returns.sample(60, random_state=31)], ignore_index=True)

# Data Dictionary
dictionary = pd.DataFrame([
    ["customers", "customer_id", "Unique customer identifier", "Should be unique; no blanks"],
    ["customers", "city", "Customer city", "Contains abbreviations, punctuation, and mixed case"],
    ["products", "list_price", "Normal selling price", "Investigate negative prices"],
    ["orders", "order_date", "Date order was placed", "Mixed string formats and bad dates included"],
    ["order_items", "quantity", "Units purchased", "Contains string flags ('MISSING') and bad numbers"],
], columns=["table", "column", "meaning", "cleaning_notes"])

# -----------------------------
# 5. Export Workbook & CSVs
# -----------------------------
path = "messy_ecommerce_data_analyst_project.xlsx"
with pd.ExcelWriter(path, engine="openpyxl") as writer:
    customers.to_excel(writer, sheet_name="RAW_customers", index=False)
    products.to_excel(writer, sheet_name="RAW_products", index=False)
    orders.to_excel(writer, sheet_name="RAW_orders", index=False)
    order_items.to_excel(writer, sheet_name="RAW_order_items", index=False)
    returns.to_excel(writer, sheet_name="RAW_returns", index=False)
    dictionary.to_excel(writer, sheet_name="Data_Dictionary", index=False)

    wb = writer.book
    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        for cell in ws[1]:
            cell.font = Font(bold=True)
        for col in range(1, ws.max_column + 1):
            max_len = min(
                max(len(str(ws.cell(row=r, column=col).value or "")) for r in range(1, min(ws.max_row, 200)+1)) + 2,
                35
            )
            ws.column_dimensions[get_column_letter(col)].width = max_len

csv_dir = "ecommerce_raw_csvs"
os.makedirs(csv_dir, exist_ok=True)
for name, df in {
    "customers.csv": customers,
    "products.csv": products,
    "orders.csv": orders,
    "order_items.csv": order_items,
    "returns.csv": returns,
}.items():
    df.to_csv(os.path.join(csv_dir, name), index=False)

print(f"File created successfully: {path}")